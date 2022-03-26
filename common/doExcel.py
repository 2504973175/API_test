
from openpyxl import load_workbook
class DoExcel:
   """读写Excel文件"""
   def get_data(self,filename,sheetname):
    wb=load_workbook(filename)
    sheet=wb[sheetname]

    test_data=[]
    #读取Excel数据存入列表
    for i in range(2,sheet.max_row+1):
        row_data={}
        row_data['case_id'] = sheet.cell(i, 1).value #第一列为case_id
        row_data['module']=sheet.cell(i,2).value
        row_data['title']=sheet.cell(i,3).value
        row_data['headers'] = sheet.cell(i, 4).value
        row_data['Cookie']=sheet.cell(i,5).value
        row_data['method'] = sheet.cell(i, 6).value
        row_data['url'] = sheet.cell(i, 7).value
        row_data['data'] = sheet.cell(i,8).value
        row_data['expected_code'] = sheet.cell(i, 9).value #预期状态码
        row_data['actual_code'] = sheet.cell(i, 10).value #实际状态码
        row_data['response']=sheet.cell(i,11).value
        row_data['auth']=sheet.cell(i,13).value
        row_data['file']=sheet.cell(i,14).value

        test_data.append((row_data))
    return test_data
   #写入Excel方法
   def write_back(self,filename,sheet_name,i,k,value):
       #i为写入的行k为写入的列
        wb=load_workbook(filename)
        sheet=wb[sheet_name]
        sheet.cell(i,k).value=value#写入表格第14列
        wb.save(filename)

if __name__ == '__main__':
    test_data=DoExcel().get_data("D:/接口实战/接口自动化用例.xlsx",'storm')
    # print(test_data)
    DoExcel().write_back("D:/接口实战/接口自动化用例.xlsx",'storm',2,5)

